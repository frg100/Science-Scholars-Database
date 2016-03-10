from tinydb import TinyDB, Query
import openpyxl
db = TinyDB('db.json')
#creates a new file named db
xl = openpyxl.load_workbook('recentmentors.xlsx')
#opens the excel spreadsheet
sheet = xl.get_sheet_by_name('Sheet1')
#makes sheet1 the main sheet

	

def separate_phone(full_name):
	parts = full_name.split('(')
	stripped_name = parts[0]
	try:
		number = '(' + parts[1]
	except IndexError:
		number = ''
	return [stripped_name, number]


for x in range(1,79):
    #iterates through each row
    full_name = sheet.cell(row = x, column = 3).value
    affiliation = sheet.cell(row = x, column = 2).value
    topic = sheet.cell(row = x, column = 4).value
    student = sheet.cell(row = x, column = 1).value
    email = sheet.cell(row = x, column = 5).value
    parts = separate_phone(full_name)
    full_name  = parts[0]
    phone = parts[1]
    print full_name
    db.insert({'Name': full_name,'Affiliation': affiliation, 'Phone': phone, 'topic': topic, 'student': student, 'Email': email})

