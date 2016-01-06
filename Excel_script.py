from tinydb import TinyDB, Query
import openpyxl
db = TinyDB('db.json')
#creates a new file named db
xl = openpyxl.load_workbook('Mentors.xlsx')
#opens the excel spreadsheet
sheet = xl.get_sheet_by_name('Sheet1')
#makes sheet1 the main sheet


def consolidate_name( prefix, first_name, middle_initial, last_name ):
	if prefix is None:
		prefix = ''
	if middle_initial is None:
		middle_initial  = ''
	if first_name is None:
		first_name = ''
	if last_name is None:
		last_name  = ''
	full_name = prefix + ' ' + first_name + ' ' + middle_initial + ' ' + last_name
	return full_name

def separate_phone(full_name, phone):
	parts = full_name.split('(')
	stripped_name = parts[0]
	try:
		number = '(' + parts[1]
	except IndexError:
		number = ''
	return [stripped_name, number]

for x in range(2,151):
    #iterates through each row
    prefix = sheet.cell(row = x, column = 1).value
    first_name = sheet.cell(row = x, column = 2).value
    middle_initial = sheet.cell(row = x, column = 3).value
    last_name = sheet.cell(row = x, column = 4).value
    affiliation = sheet.cell(row = x, column = 6).value
    phone  = sheet.cell(row = x, column = 7).value
    topic = sheet.cell(row = x, column = 8).value
    student = sheet.cell(row = x, column = 9).value
    full_name = consolidate_name(prefix,first_name,middle_initial,last_name)
    db.insert({'Name': full_name,'Affiliation': affiliation, 'Phone': phone, 'topic': topic, 'student': student})

for x in range(151,376):
    #iterates through each row
    full_name = sheet.cell(row = x, column = 5).value
    affiliation = sheet.cell(row = x, column = 6).value
    topic = sheet.cell(row = x, column = 8).value
    student = sheet.cell(row = x, column = 9).value
    parts = separate_phone(full_name,phone)
    full_name  = parts[0]
    phone = parts[1]
    db.insert({'Name': full_name,'Affiliation': affiliation, 'Phone': phone, 'topic': topic, 'student': student})

