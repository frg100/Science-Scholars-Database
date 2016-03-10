from tinydb import TinyDB, Query
from re import search
db = TinyDB('db.json')

print '''
-------------------------------------------------------------------------
|WELCOME TO THE EHS SCIENCE SCHOLARS MENTOR DATABASE                     |
|                                                                        |
|DESIGNED BY FEDERICO REYES AND THE E-TECH CLUB 2015                     |
|                                                                        |
|THIS DATABASE CAN BE USED TO SEARCH FOR AND ADD MENTORS TO THE DATABASE |
-------------------------------------------------------------------------
'''

def consolidate_name( prefix, first_name, middle_initial, last_name ):
	if prefix is None:
		prefix = ''
	if middle_initial is None:
		middle_initial  = ''
	if first_name is None:
		first_name = ''
	if last_name is None:
		last_name  = ''
	full_name = prefix + ' ' + first_name + ' ' + middle_initial + ' ' + last_name

def Selection():
	a = 0
	while a != 1:
		selection = raw_input('''
-------------------------------------------------
|Type 0 if you would like to SEARCH the DATABASE|
|                                               |
|Type 1 if you would like to ADD to the DATABASE|
|                                               |
|Type 2 if you would like to EXIT the program   |
------------------------------------------------|

Make Your Selection : ''')

		if selection == '0':
			search()
			a = 1

		elif selection == '1':
			ynexcel = raw_input("\nIf you would you like to add using an excel sheet in the proper format, press 0\nIf you would like to add the data manually, press 1")
			if ynexcel == '0':
				add_excel()
				a = 1
			elif ynexcel == '1':
				add()
				a = 1
			else:
				pass

		elif selection == '2':
			print "\n"
			exit()

		else:
			a = 0
			print "\nThis is an invalid selection, please try again..."

def search():
	mentor = Query()
	d = 0
	while d != 1:

		print ('''
----------------------------------------------------
|What would you like to search by?                  |
|                                                   |
|If you would like to search by FIRST NAME, type 0  |
|                                                   |
|If you would like to search by LAST NAME, type 1   |
|                                                   |
|If you would like to search by AFFILIATION, type 2 |
|                                                   |
|If you would like to search by TOPIC, type 3       |
|                                                   |
|If you would like to search by STUDENT, type 4     |
|                                                   |
|TIP: Less is more!                                 |
|   (If the affiliation name is "Columbia",         |
|   a search of "Columbia" will give you a result   |
|   while a search of "Columbia University" won't)  |
---------------------------------------------------
''')
		selection = raw_input('\n Please Make Your Selection : ')
		if selection == '0':
			d = 1
			query = 'First_name'

		elif selection == '1':
			d = 1
			query = 'Last_name'

		elif selection == '2':
			d = 1
			query = 'Affiliation'

		elif selection == '3':
			d = 1
			query = 'topic'

		elif selection == '4':
			d = 1
			query = 'student'
		else:
			d = 0
			print "This is an invalid selection, please try again..."
	user_input = raw_input("\nPlease enter search text : ")
	print '''
-----------------------------------------------------------------------------------------------------------------------
|                                                                                                                      |
|                                                                                                                      |
|                                                          RESULTS                                                     |
|                                                                                                                      |
|                                                                                                                      |
-----------------------------------------------------------------------------------------------------------------------
	'''
	for x in db.all():
		try:
			low = x[query].lower()
			spl = low.split()
			mutatedstring = ''.join(spl)
			if user_input.lower() in mutatedstring:
				print ( '''
---------------------------------------------
|Mentor name : %s                           
|
|Topic : %s
|
|Affiliation : %s
|
|Phone Number : %s
|
|Student : %s
|
|Email: %s
|
|Year: %s
---------------------------------------------
''' %(x['Name'], x['topic'], x['Affiliation'], x['Phone'], x['student'], x['Email'], x['Year'] ))
		except TypeError:
			pass
		except AttributeError:
			pass
		except KeyError:
			print ( '''
---------------------------------------------
|Mentor name : %s                           
|
|Topic : %s
|
|Affiliation : %s
|
|Phone Number : %s
|
|Student : %s
---------------------------------------------
''' %(x['Name'], x['topic'], x['Affiliation'], x['Phone'], x['student']))
	#print db.search(mentor.query.search('%s' %user_input))
	#info = db.get(mentor.query == '%s' %user_input)
	'''
	try:
		for i in info:
			print ("%s : %s") %(i, info[i])
		hi = raw_input("Press enter to continue: ")
	except TypeError:
		print ("There are no results to match that query...")
	'''
	Selection()

def consolidate_name( prefix, first_name, middle_initial, last_name ):
	if prefix is None:
		prefix = ''
	if middle_initial is None:
		middle_initial  = ''
	if first_name is None:
		first_name = ''
	if last_name is None:
		last_name  = ''
	full_name = prefix + ' ' + first_name + ' ' + middle_initial + ' ' + last_name
	return full_name

def add_excel():
	print "Make sure the excel sheet is in the proper format. Also make sure it is in the application folder"
	import openpyxl
	excel_sheet = raw_input('''What is the name of the excel sheet?\nFor example, you could name your sheet "mentors.xlsx"\nMake sure to type in the exact name: ''')
	xl = openpyxl.load_workbook(excel_sheet)
	sheet = xl.get_sheet_by_name('Sheet1')
	number_of_rows = int(raw_input("What is the ROW number of the last mentor in the sheet?: "))
	for x in range(3,number_of_rows+1):
		year = sheet.cell(row = 1, column = 2).value
    	prefix = sheet.cell(row = x, column = 1).value
    	first_name = sheet.cell(row = x, column = 2).value
    	middle_initial = sheet.cell(row = x, column = 3).value
    	last_name = sheet.cell(row = x, column = 4).value
    	affiliation = sheet.cell(row = x, column = 6).value
    	phone  = sheet.cell(row = x, column = 7).value
    	topic = sheet.cell(row = x, column = 8).value
    	student = sheet.cell(row = x, column = 9).value
    	full_name = consolidate_name(prefix,first_name,middle_initial,last_name)
    	db.insert({'Name': full_name,'Affiliation': affiliation, 'Phone': phone, 'topic': topic, 'student': student, 'Year': year})


def add():
	a = 0
	while a != 1:
		print ("\nPlease enter the information as prompted")
		print ("\n NOTE: If information not available, press enter to skip... \n")
		print"----------------------------------------------------------------------------"
		prefix = raw_input("\nPrefix: ")
		first_name = raw_input("\nFirst Name: ")
		middle_initial = raw_input("\nMiddle Initial: ")
		last_name = raw_input("\nLast Name: ")
		affiliation = raw_input("\nAffiliation: ")
		phone  = raw_input("\nPhone Number: ")
		topic = raw_input("\nTopic: ")
		student = raw_input("\nStudent: ")
		year = raw_input("\nYear: ")
		print ("\n\nHere is what you inputted: \n\n    Scientist: %s %s %s %s \n\n    Affiliation: %s \n\n    Phone Number: %s \n\n    Topic: %s \n\n    Student: %s \n\n    Year: %s \n\n") %(prefix, first_name, middle_initial, last_name, affiliation, phone, topic, student, year)
		selection = raw_input('''
Is this information ok? 

-------------------------------------------------------------------------------
|If you would like to confirm this information and continue, type 0            |
|                                                                              |
|If you would like to input the information again, type 1                      |
|                                                                              |
|If you would like to exit, type 2                                             |
------------------------------------------------------------------------------


Please Make Your Selection : \n
''')
		b = 0
		while b != 1:
			if selection == '0':
				a = 1
				b = 1
				full_name = consolidate_name(prefix,first_name,middle_initial,last_name)
				db.insert({'Name': full_name,'Affiliation': affiliation, 'Phone': phone, 'topic': topic, 'student': student, 'Year': year})

			elif selection == '1':
				a = 0
				b = 1
			elif selection == '2':
				a = 1
				b = 1

			else:
				a = 0
				b = 0
				print "This is an invalid selection, please try again..."
	Selection()

def main():
	Selection()

main()